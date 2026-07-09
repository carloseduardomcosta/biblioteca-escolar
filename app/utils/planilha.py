"""Leitura robusta de planilhas para importação (compartilhado por livros e alunos).

Aceita .xlsx (Excel) e CSV, detectando a codificação do CSV (UTF-8/BOM/latin-1)
para não corromper acentos. Colunas são normalizadas para minúsculas.
"""
import csv
import io


def parse_ano(valor):
    """Extrai um ano como int. Aceita '2020', '2020.0', '  2020 '. None se vazio/inválido."""
    s = str(valor or '').strip()
    if not s:
        return None
    if s.endswith('.0'):
        s = s[:-2]
    return int(s) if s.isdigit() else None


def linhas_planilha(file):
    """Gera dicts (colunas em minúsculas) a partir de um .xlsx OU CSV."""
    nome = (getattr(file, 'filename', '') or '').lower()

    if nome.endswith(('.xlsx', '.xlsm')):
        from openpyxl import load_workbook
        wb = load_workbook(file.stream, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            return
        cols = [str(h).strip().lower() if h is not None else '' for h in header]
        for r in it:
            if r is None or all(v is None for v in r):
                continue
            valores = ['' if v is None else str(v).strip() for v in r]
            valores += [''] * (len(cols) - len(valores))
            yield dict(zip(cols, valores))
        return

    # CSV — tenta várias codificações, sem perder acento
    raw = file.stream.read()
    if isinstance(raw, str):
        texto = raw
    else:
        for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                texto = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            texto = raw.decode('latin-1', errors='replace')
    reader = csv.DictReader(io.StringIO(texto))
    for row in reader:
        yield {
            (k.strip().lower() if isinstance(k, str) else k):
            (v.strip() if isinstance(v, str) else ('' if v is None else v))
            for k, v in row.items()
        }
